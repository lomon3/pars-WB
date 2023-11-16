import os
import glob
import openpyxl
import asyncio
from asyncio import sleep
import aiohttp
import json

async def get_brands(art_list):
    result = set()
    url = f"https://card.wb.ru/cards/detail?spp=10&reg=1&locale=ru&dest=-1216601,-115136,-421732,123585595&nm={';'.join(art_list)}"
    
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get(url) as response:
                data = await response.json()
        except Exception as e:
            print(f"Failed to get data for art= {art_list}: {e}")
            return result

    products = data["data"]["products"]
    for product in products:
        if product:
            result.add(product["brandId"])

    return result

async def get_brand_filter_and_total(keyword, fbrand=None):
    url = f"https://search.wb.ru/exactmatch/ru/male/v4/search?appType=1&curr=rub&dest=-1257786&query={keyword}&resultset=filters"
    if fbrand:
        #print(f"Получили список {fbrand}, теперь запрос {url}")
        url += f"&fbrand={fbrand}"
    
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get(url) as response:
                content_type = response.headers.get('Content-Type')
                if 'application/json' in content_type:
                    data = await response.json()
                else:
                    # Пробуем декодировать содержимое как JSON, даже если тип содержимого не соответствует
                    text = await response.text()
                    data = json.loads(text)
                
                # Проверяем наличие ключа 'data' в ответе
                if 'data' not in data:
                    print(f"Ответ не содержит ожидаемых данных: {data}")
                    return [], 0
                
        except Exception as e:
            print(f"Failed to decode JSON for keyword= {keyword}: {e}")
            return [], 0

    brand_items = next((filter_["items"] for filter_ in data["data"]["filters"] if filter_["name"] == "Бренд"), [])
    total_pages = (data["data"]["total"] + 99) // 100  # Округление вверх
    
    if fbrand:
        additional_info = ", но мы проверили только 30, " if total_pages > 30 else ""
        print(f"С фильтром по бренду всего страниц {total_pages}, и там {data['data']['total']} товаров{additional_info}")
    else:
        print(f"Без фильтров всего страниц {total_pages}, и там {data['data']['total']} товаров")

    return [item["id"] for item in brand_items], total_pages

async def fetch_data(session, url, retries=2, delay=4):
    for attempt in range(1, retries+1):
        try:
            async with session.get(url) as response:
                if response.status == 429:  # Too Many Requests
                    print(f"Попытка {attempt}. Статус {response.status} по {url}")
                    await sleep(delay)
                    continue
                if response.status != 200:
                    print(f"Попытка {attempt}. Error {response.status} when accessing {url}")
                    return None

                content_type = response.headers.get('Content-Type')
                if 'application/json' in content_type:
                    data = await response.json()
                    if 'data' not in data:
                        print(f"Попытка {attempt}. Unexpected response structure from {url}: {data}")
                        return None
                    elif attempt != 1:
                        print(f"Попытка {attempt} была успешной")
                    return data
                else:
                    # Пробуем декодировать содержимое как JSON, даже если тип содержимого не соответствует
                    text = await response.text()
                    try:
                        data = json.loads(text)
                        if 'data' not in data:
                            print(f"Попытка {attempt}. Unexpected response structure from {url}: {data}")
                            return None
                        elif attempt != 1:
                            print(f"Попытка {attempt} была успешной")
                        return data
                    except json.JSONDecodeError:
                        print(f"Попытка {attempt}. Unexpected content type '{content_type}' received from {url}: {text}")
                        return None
        except Exception as e:
            print(f"Failed to fetch data from {url}: {e}")
            return None
    return None  # Если все попытки закончились неудачей


async def get_positions_by_keyword(session, keyword, art_list, total_pages, fbrand=None, max_pages_without_brand=40):
    position_dict = {art: None for art in art_list}
    #max_pages_without_brand = 40  # Ограничение на количество страниц без фильтра бренда
    page = 1
    
    # Сначала поиск без фильтра бренда
    while None in position_dict.values() and page <= min(max_pages_without_brand, total_pages):
        #print(f"список брендов: {fbrand != None} и кол-во страниц {total_pages} и ограничитель {max_pages_without_brand}")
        url = f"https://search.wb.ru/exactmatch/ru/common/v4/search?page={page}&appType=1&curr=rub&dest=-1257786&lang=ru&locale=ru&query={keyword}&resultset=catalog"
        data = await fetch_data(session, url)
        if data and 'data' in data:
            products = data["data"]["products"]
            for art in position_dict:
                if position_dict[art] is None:
                    index = next((index for index, prod in enumerate(products) if prod["id"] == int(art)), None)
                    if index is not None:
                        position_dict[art] = ((page-1)*100)+index+1
        page += 1

    # Если еще остались не найденные товары, применяем фильтр бренда и начинаем поиск с первой страницы
    if None in position_dict.values():
        if not fbrand:
            brand_list = await get_brands(art_list)  # Получаем список брендов
            fbrand = '%3B'.join(str(brand) for brand in brand_list)
            _, total_pages = await get_brand_filter_and_total(keyword, fbrand=fbrand)  # Обновляем total_pages с фильтром бренда
        
        max_limit_with_brand = 30
        page = 1
        while None in position_dict.values() and page <= min(total_pages, max_limit_with_brand):
            #print(f"список брендов: {fbrand != None} и кол-во страниц {total_pages} и ограничитель {max_pages_without_brand}")
            url = f"https://search.wb.ru/exactmatch/ru/common/v4/search?page={page}&appType=1&curr=rub&dest=-1257786&lang=ru&locale=ru&query={keyword}&resultset=catalog&fbrand={fbrand}"
            data = await fetch_data(session, url)
            if data and 'data' in data:
                products = data["data"]["products"]
                for art in position_dict:
                    if position_dict[art] is None:
                        index = next((index for index, prod in enumerate(products) if prod["id"] == int(art)), None)
                        if index is not None:
                            position_dict[art] = f"{max_pages_without_brand}00+"
            page += 1

    return position_dict, total_pages

async def pars_selected_files(selected_file, max_pages_without_brand):
    print(f"Начало обработки файла: {selected_file}")
    wb = openpyxl.load_workbook(selected_file)
    sheet = wb['Общий отчет']
    max_col = sheet.max_column

    art_list = set()
    for col in range(4, max_col + 1):
        art = str(sheet.cell(row=2, column=col).value)
        art_list.add(art)
    art_list.discard('None')

    brand_list = await get_brands(art_list)
    print(f"Получены ID брендов: {brand_list}")
    fbrand = '%3B'.join(str(brand) for brand in brand_list)

    for row in sheet.iter_rows(min_row=3, min_col=4, max_row=sheet.max_row, max_col=max_col):
        for cell in row:
            cell.value = ""
    print('Лист очищен')

    keywords = {row[0].row: str(row[1].value) for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row)}

    async with aiohttp.ClientSession() as session:
        for index, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet.max_row), start=3):
            keyword = keywords[index]
            print(f"Начало обработки ключевого слова: {keyword}")
            brand_ids, total_pages = await get_brand_filter_and_total(keyword)
            #print(f"total_pages: {total_pages}")

            position_dict, total_pages = await get_positions_by_keyword(session, keyword, art_list, total_pages, None, max_pages_without_brand)
            position_dict_with_brands, _ = await get_positions_by_keyword(session, keyword, art_list, total_pages, fbrand, max_pages_without_brand)

            #for key in position_dict:
                #if position_dict_with_brands.get(key) is not None:
                    #position_dict[key] = "1000+"

            for col in range(4, max_col + 1):
                if str(sheet.cell(row=2, column=col).value) in position_dict:
                    sheet.cell(row=row[0].row, column=col, value=position_dict[str(sheet.cell(row=2, column=col).value)])
            print(f"Ключевое слово {keyword} обработано!")
    success = False
    while not success:
        try:
            wb.save(selected_file)
            print(f"Файл {selected_file} обработан и сохранен!")
            success = True
        
        except PermissionError as e:
            print(f"Ошибка доступа к файлу {selected_file}. Пожалуйста, закройте файл и нажмите Enter для повторной попытки, или введите 'ggwp' для отмены. Ошибка: {e}")
            response = input().strip().lower()
            if response == 'ggwp':
                print("Операция отменена пользователем.")
                break

        except Exception as e:
            print(f"Произошла неизвестная ошибка при обработке файла {selected_file}: {e}")
            break



def select_file():
    # Получение текущего рабочего каталога
    current_dir = os.getcwd()

    # Шаблон имени файла для поиска (например, '*.xlsx')
    file_pattern = '*.xlsx'

    # Полный путь к папке с файлами Excel
    folder_path = os.path.join(current_dir, file_pattern)

    # Получение списка файлов Excel в папке
    excel_files = glob.glob(folder_path)

    # Получение только имен файлов без путей
    excel_file_names = [os.path.basename(file) for file in excel_files]

    # Вывод списка файлов с порядковыми номерами
    for i, file_name in enumerate(excel_file_names):
        print(f"{i+1}. {file_name}")
    print("0. обработать все файлы")
    print()

    # Пользователь выбирает номер файла для обработки
    selected_file_num = int(input("Введите номер файла для обработки: "))
    print()

    # Проверка выбранного номера файла
    if 1 <= selected_file_num <= len(excel_files):
        selected_file_name = excel_file_names[selected_file_num - 1]
        # Здесь можете обработать выбранный файл
        print(f"Выбран файл: {selected_file_name}")
        print()
        return [str(selected_file_name)]
    else:
        print("Выбраны все файлы")
        return excel_file_names

async def main():
    start_stop = "start"
    while start_stop != "stop":
        try:
            selected_files = []
            selected_files += select_file()
            max_pages_without_brand = int(input("Введите максимальное количество страниц без фильтра бренда (по умолчанию 40): ") or 40)
            for file in selected_files:
                await pars_selected_files(file, max_pages_without_brand)
            start_stop = "stop" if input("Введите: 1 - выход; Enter - выбрать другой файл: ") == "1" else start_stop
        except Exception as e:
            print(f"Произошла ошибка: {e}")
            input("Нажмите Enter, чтобы продолжить...")
        print()
    
    print("Made by https://t.me/ArChernushevich")
    input("Нажмите Enter, чтобы выйти...")
asyncio.run(main())
