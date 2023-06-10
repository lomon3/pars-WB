
# pip install pandas
# pip install requests
# pip install pyinstaller

import requests
import openpyxl

# Открыть книгу Excel и получить листы
wb = openpyxl.load_workbook('position_v2.xlsx')
sheet = wb['Общий отчет']

# группируем список
unique_brands = set()
chunk_size = 300
art_list = []
for row in range(2, sheet.max_row + 1):
    art = str(sheet.cell(row=row, column=1).value)
    art_list.append(art)
    article_chunks = [art_list[i:i+chunk_size]
                      for i in range(0, len(art_list), chunk_size)]

for chunk in article_chunks:
    url = f"https://card.wb.ru/cards/detail?spp=10&reg=1&locale=ru&dest=-1216601,-115136,-421732,123585595&nm={';'.join(chunk)}"
    # Отправляем запрос и получаем ответ в формате JSON
    try:
        response = requests.get(url)
        data = response.json()
    except Exception as e:
        print(f"Failed to get data for art= {chunk}: {e}")
        continue
    products = data["data"]["products"]
    for product in products:
        if product:
            unique_brands.add(product["brandId"])
        else:
            brandId = ""

for brand in unique_brands:

    for col in range(2, sheet.max_column + 1):
        keyword = str(sheet.cell(row=1, column=col).value)
        page = 1
        empty = "no empty"
        while empty != "empty":

            url = f"https://search.wb.ru/exactmatch/ru/common/v4/search?page={page}&appType=1&curr=rub&dest=-1257786&lang=ru&locale=ru&query={keyword}&resultset=catalog&fbrand={brand}"

            # Отправляем запрос и получаем ответ в формате JSON
            try:
                response = requests.get(url)
                data = response.json()
            except Exception as e:
                print(
                    f"Failed to get data for keyword= {keyword} page= {page}: {e}")

            products = data["data"]["products"]

            if products:
                for product in products:
                    if product:
                        for row in range(2, sheet.max_row + 1):
                            if product["id"] == (sheet.cell(row=row, column=1).value):
                                # print(f"{product['id']} match {str(sheet.cell(row=row, column=1).value)}")
                                sheet.cell(row=row, column=col, value="V")
            else:
                empty = "empty"

            print(f" {brand} {keyword} - page #{page} {empty}")
            page += 1


# Сохраняем excel-файл
wb.save("position_v2.xlsx")
print("Made by https://t.me/ArChernushevich")
input("Нажмите Enter, чтобы выйти...")
