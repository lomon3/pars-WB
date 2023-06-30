import requests
import openpyxl
import os
import glob


def select_file():
    # Получение текущего рабочего каталога
    current_dir = os.getcwd()

    # Шаблон имени файла для поиска (например, '*.xlsx')
    file_pattern = '*.xlsx'

    # Полный путь к папке с файлами Excel
    folder_path = os.path.join(current_dir, file_pattern)

    # Получение списка файлов Excel в папке
    excel_files = glob.glob(folder_path)

    # Получение только имен файлов без путей
    excel_file_names = [os.path.basename(file) for file in excel_files]

    # Вывод списка файлов с порядковыми номерами
    for i, file_name in enumerate(excel_file_names):
        print(f"{i+1}. {file_name}")
    print("0. обработать все файлы")
    print()

    # Пользователь выбирает номер файла для обработки
    selected_file_num = int(input("Введите номер файла для обработки: "))
    print()

    # Проверка выбранного номера файла
    if 1 <= selected_file_num <= len(excel_files):
        selected_file_name = excel_file_names[selected_file_num - 1]
        # Здесь можете обработать выбранный файл
        print(f"Выбран файл: {selected_file_name}")
        print()
        return [str(selected_file_name)]
    else:
        print("Выбраны все файлы")
        return excel_file_names

def get_brands(art_list):
    result = set()
    url = f"https://card.wb.ru/cards/detail?spp=10&reg=1&locale=ru&dest=-1216601,-115136,-421732,123585595&nm={';'.join(art_list)}"
    # Отправляем запрос и получаем ответ в формате JSON
    try:
        response = requests.get(url)
        data = response.json()
    except Exception as e:
        print(f"Failed to get data for art= {art_list}: {e}")

    products = data["data"]["products"]
    for product in products:
        if product:
            result.add(product["brandId"])
        else:
            brandId = ""
    
    return result

def pars_selected_files(selected_file):
    print(selected_file)
    # Открыть книгу Excel и получить листы
    wb = openpyxl.load_workbook(selected_file)
    sheet = wb['Общий отчет']
    max_col = sheet.max_column

    art_list = set()
    for col in range(4, max_col + 1):
      art = str(sheet.cell(row=2, column=col).value)
      art_list.add(art)
    art_list.discard('None')

    brand_list = get_brands(art_list)
    print(f"get brand_ids: {brand_list}")
    fbrand = '%3B'.join(str(brand) for brand in brand_list)

    for row in sheet.iter_rows(min_row=3, min_col=4, max_row=sheet.max_row, max_col=max_col):
        for cell in row:
            cell.value = ""
    print('sheet cleaned')

    # Индексированный словарь для хранения значений столбца 2
    keywords = {row[0].row: str(row[1].value) for row in sheet.iter_rows(
        min_row=3, max_row=sheet.max_row)}
    
    for index, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet.max_row), start=3):
        # Создание словаря для хранения позиций артикулов на странице поиска
        position_dict = {art: None for art in art_list}
        keyword = keywords[index]
        page = 1
        empty = "no empty"

        while empty != "empty" and page <= 10 and None in position_dict.values():
            url = f"https://search.wb.ru/exactmatch/ru/common/v4/search?page={page}&appType=1&curr=rub&dest=-1257786&lang=ru&locale=ru&query={keyword}&resultset=catalog"
            try:
                response = requests.get(url)
                data = response.json()

                if 'data' in data:
                    products = data["data"]["products"]
                    if products:
                        for art in position_dict:
                            if position_dict[art] is None:
                                index = next((index for index, prod in enumerate(products) if prod["id"] == int(art)), None)
                                if index is not None:
                                    position_dict[art] = ((page-1)*100)+index+1
                    else:
                        empty = "empty"
                else:
                    empty = "empty"

            except Exception as e:
                print(f"Failed to get data for keyword = {keyword} page = {page}: {e}")

            page += 1

            if page < 11:
                print(f"{keyword} - page #{page-1} {empty}")
            else:
                page = 1
                while empty != "empty" and page <= 25 and None in position_dict.values():
                    url2 = f"https://search.wb.ru/exactmatch/ru/common/v4/search?page={page}&appType=1&curr=rub&dest=-1257786&lang=ru&locale=ru&query={keyword}&resultset=catalog&fbrand={fbrand}"
                    try:
                        response = requests.get(url2)
                        data = response.json()

                        if 'data' in data:
                            products = data["data"]["products"]
                            if products:
                                for art in position_dict:
                                    if position_dict[art] is None:
                                        index = next((index for index, prod in enumerate(products) if prod["id"] == int(art)), None)
                                        if index is not None:
                                            position_dict[art] = "1000+"
                            else:
                                empty = "empty"
                        else:
                            empty = "empty"

                    except Exception as e:
                        print(f"Failed to get data for keyword = {keyword} page = {page}: {e}")
                    page += 1

                print(f"{keyword} - Checked {page-1} pages with a filter by brand, page {page} is {empty}")
                

            
        for col in range(4, max_col + 1):
            if str(sheet.cell(row=2, column=col).value) in position_dict:
                sheet.cell(row=row[0].row, column=col, value=position_dict[str(sheet.cell(row=2, column=col).value)])


    # Сохраняем excel-файл
    wb.save(selected_file)
    print()

start_stop = "start"
while start_stop != "stop":
    selected_files = []
    selected_files += select_file()
    for files in selected_files:
        pars_selected_files(files)
    start_stop = "stop" if input("Введите: 1 - выход; Enter - выбрать другой файл: ") == "1" else start_stop
    print()
    

print("Made by https://t.me/ArChernushevich")
input("Нажмите Enter, чтобы выйти...")
