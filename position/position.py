
# pip install pandas
# pip install requests
# pip install pyinstaller

import requests
import openpyxl


# Открыть книгу Excel и получить листы
wb = openpyxl.load_workbook('position.xlsx')
sheet = wb['Общий отчет']

for col in range(2, sheet.max_column + 1):
    keyword = str(sheet.cell(row=1, column=col).value)
    page = 1
    empty = "no empty"
    while empty != "empty":

        url = f"https://search.wb.ru/exactmatch/ru/common/v4/search?page={page}&appType=1&curr=rub&dest=-1257786&lang=ru&locale=ru&query={keyword}&resultset=catalog&fbrand=121380"

        # Отправляем запрос и получаем ответ в формате JSON
        try:
            response = requests.get(url)
            data = response.json()
        except Exception as e:
            print(f"Failed to get data for keyword= {keyword} page= {page}: {e}")
        
        products = data["data"]["products"]

        if products:
            for product in products:
                if product:
                    for row in range(2, sheet.max_row + 1):
                        if product["id"] == (sheet.cell(row=row, column=1).value):
                            #print(f"{product['id']} match {str(sheet.cell(row=row, column=1).value)}")
                            sheet.cell(row=row, column=col, value="V")
        else:
            empty = "empty"

        print(f" {keyword} - page #{page} {empty}")
        page += 1


# Сохраняем excel-файл
wb.save("position.xlsx")
input("Нажмите Enter, чтобы выйти...")
