
# pip install pandas
# pip install requests
# pip install pyinstaller

import requests
import openpyxl
import requests

# Получаем значение spp
def get_spp():
    headers = {
        'Cookie': "_wbauid=2725866931671562445; ___wbu=4d7fb948-d2d6-4346-bdaf-24c4aaf642a5.1671562445; BasketUID=fc9ac6dd-d0a3-4eb9-907b-826afdebee8f; WILDAUTHNEW_V3=B522E25B0D2D22627383FE6A5BC3835E2A07AC77CEFD60DAB8B50910DC553AAF7440B66261D6877F35A55A10BFBDCB36B0021A8F62672BFE1E7B7DCE459212D05968AD4B87934DD0F00FC876A86F15379F318C1A092F9227D57BF90912E8B4779F258FCCFAF2610B66FACB854A0CF498492F52D10E1A91A8AEFD99FB39B6546E3E37F31003BD5317445446D9F4255E12FA7DE9B4B7A73ADAA9C2904992E952AD494A2382A6324E8533CAAB3DDCCD80F3F6FD813D2BF20742AE5B85EB8C401D9E8BF39885493F5B82C1C65094F1EA924CA11B0589413D5F83079BA0EA98AE5546C80B61E159176C942B8669BB267AE50C0315F518094653817BC087C1D3E4625AAB3F49A42D80B3EFC848B311A9FA20AB39C9A09D76E3BDE52C441058AEFD01A4ABC2EB703953D7A5BC32FE2636DA2174580CF360; external-locale=ru; x-supplier-id-external=d99a1f06-4d18-5d70-8ee4-d0fefc8a193f; um=uid%3Dw7TDssOkw7PCu8KwwrPCs8KywrjCucKzwrPCuA%253d%253d%3Aproc%3D100%3Aehash%3Dd41d8cd98f00b204e9800998ecf8427e; __wba_s=1; wbx-validation-key=6efd7552-485b-4ba8-899d-fce2056c2352; ___wbs=ed012f2f-f441-40e5-bed5-729290e7bee0.1693806859",
    }

    url = 'https://www.wildberries.ru/webapi/personalinfo/extrainfo'

    while True:
        response = requests.post(url, headers=headers)
        if response.status_code == 200:
            personal_discount = response.json()['value']['discountInfo']['personalDiscount']
            return personal_discount
        else:
            print(f"Ошибка запроса: {response.status_code}. Повторная попытка...")

spp = input("Введите СПП или оставьте пустым для автоматической загрузки СПП: ") or get_spp()
print(f'Сегодня СПП = {spp}%')

# Открыть книгу Excel и получить листы
wb = openpyxl.load_workbook('Отчет по скидкам.xlsx')
old_sheet = wb['Общий отчет']
sheet = wb.create_sheet('pars')

# Пройти по каждой строке на исходном листе
for row in old_sheet.iter_rows(min_row=1, max_col=5, max_row=old_sheet.max_row):
    # Проверить значение ячейки в первом столбце
    if row[0].value in ["WiMi", "BASEUS", "Бренд"]:
        # Создать новый объект строки и заполнить его значениями из соответствующих ячеек
        new_row = []
        for cell in row:
            new_row.append(cell.value)
        # Добавить строку на новый лист
        sheet.append(new_row)

# Сохранить книгу Excel
wb.save('Отчет по скидкам.xlsx')
print("Filtered brands to new sheet")

# Добавляем новые столбцы
sheet.cell(row=1, column=1, value="Арт ВБ")
sheet.cell(row=1, column=2, value="Название")
sheet.cell(row=1, column=3, value="Цена до скидок")
sheet.cell(row=1, column=4, value="Скидка поставщика")
sheet.cell(row=1, column=5, value="Цена после скидки поставщика")
sheet.cell(row=1, column=6, value="СПП")
sheet.cell(row=1, column=7, value="Цена после СПП")
sheet.cell(row=1, column=8, value="сток")
sheet.cell(row=1, column=9, value="рейт")
sheet.cell(row=1, column=10, value="отзывы")

# группируем список
chunk_size = 300
art_list = []
for row in range(2, sheet.max_row + 1):
    art = str(sheet.cell(row=row, column=5).value)
    art_list.append(art)
    article_chunks = [art_list[i:i+chunk_size]
                      for i in range(0, len(art_list), chunk_size)]

row = 2
for chunk in article_chunks:
    url = f"https://card.wb.ru/cards/detail?spp={spp}&reg=1&locale=ru&dest=-1216601,-115136,-421732,123585595&nm={';'.join(chunk)}"
    # print(url)

    # Отправляем запрос и получаем ответ в формате JSON
    try:
        response = requests.get(url)
        data = response.json()
    except Exception as e:
        print(f"Failed to get data for art= {chunk}: {e}")
        continue
    products = data["data"]["products"]
    for product in products:
        if product:

            sku = product["id"]
            name = product["name"]

            price_u = product.get("priceU")
            if price_u is not None:
                price_u = price_u / 100

            sizes = product.get("sizes", [])
            qty = sum(stock["qty"]
                      for size in sizes for stock in size.get("stocks", []))

            rating = product["rating"]
            feedbacks = product["feedbacks"]

            extended = product.get("extended")
            if extended is not None:
                basic_sale = extended.get("basicSale", 0)

                basic_price_u = extended.get("basicPriceU")
                if basic_price_u is not None:
                    basic_price_u = basic_price_u / 100
                else:
                    basic_price_u = price_u

                client_sale = extended.get("clientSale", 0)

                client_price_u = extended.get("clientPriceU")
                if client_price_u is not None:
                    client_price_u = client_price_u / 100
                else:
                    client_price_u = basic_price_u
            else:
                basic_sale = ""
                basic_price_u = ""
                client_sale = ""
                client_price_u = ""

        else:
            sku = ""
            name = ""
            price_u = ""
            basic_sale = ""
            basic_price_u = ""
            client_sale = ""
            client_price_u = ""
            qty = ""
            rating = ""
            feedbacks = ""

        # Записываем результаты в excel-файл
        row_values = [sku, name, price_u, basic_sale, basic_price_u,
                      client_sale, client_price_u, qty, rating, feedbacks]
        for i, value in enumerate(row_values):
            sheet.cell(row=row, column=i+1, value=value)

        print(f"\rSKU #{row-1} {sku} ok", end="")
        row += 1

# Сохраняем excel-файл
wb.save("Отчет по скидкам.xlsx")
print()
print("Made by https://t.me/ArChernushevich")
input("Нажмите Enter, чтобы выйти...")
