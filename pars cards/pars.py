
# pip install pandas
# pip install requests
# pip install pyinstaller

import requests
import openpyxl

# Получаем значение spp
spp = 20

try:
    response = requests.get("http://10.178.2.65/spp/")
    spp = response.json()["spp"]
except:
    print("Failed to get spp from server")

# Открыть книгу Excel и получить листы
wb = openpyxl.load_workbook('Отчет по скидкам.xlsx')
old_sheet = wb['Общий отчет']
sheet = wb.create_sheet('pars')

# Пройти по каждой строке на исходном листе
for row in old_sheet.iter_rows(min_row=1, max_col=5, max_row=old_sheet.max_row):
    # Проверить значение ячейки в первом столбце
    if row[0].value in ["WiMi", "BASEUS", "Бренд"]:
        # Создать новый объект строки и заполнить его значениями из соответствующих ячеек
        new_row = []
        for cell in row:
            new_row.append(cell.value)
        # Добавить строку на новый лист
        sheet.append(new_row)

# Сохранить книгу Excel
wb.save('Отчет по скидкам.xlsx')
print("Filtered brands to new sheet")

# Добавляем новые столбцы
sheet.cell(row=1, column=1, value="Арт ВБ")
sheet.cell(row=1, column=2, value="Название")
sheet.cell(row=1, column=3, value="Цена до скидок")
sheet.cell(row=1, column=4, value="Скидка поставщика")
sheet.cell(row=1, column=5, value="Цена после скидки поставщика")
sheet.cell(row=1, column=6, value="СПП")
sheet.cell(row=1, column=7, value="Цена после СПП")
sheet.cell(row=1, column=8, value="сток")
sheet.cell(row=1, column=9, value="рейт")
sheet.cell(row=1, column=10, value="отзывы")

# группируем список
chunk_size = 300
art_list = []
for row in range(2, sheet.max_row + 1):
    art = str(sheet.cell(row=row, column=5).value)
    art_list.append(art)
    article_chunks = [art_list[i:i+chunk_size]
                      for i in range(0, len(art_list), chunk_size)]

row = 2
for chunk in article_chunks:
    url = f"https://card.wb.ru/cards/detail?spp={spp}&reg=1&locale=ru&dest=-1216601,-115136,-421732,123585595&nm={';'.join(chunk)}"
    # print(url)

    # Отправляем запрос и получаем ответ в формате JSON
    try:
        response = requests.get(url)
        data = response.json()
    except Exception as e:
        print(f"Failed to get data for art= {chunk}: {e}")
        continue
    products = data["data"]["products"]
    for product in products:
        if product:

            sku = product["id"]
            name = product["name"]

            price_u = product.get("priceU")
            if price_u is not None:
                price_u = price_u / 100

            sizes = product.get("sizes", [])
            qty = sum(stock["qty"]
                      for size in sizes for stock in size.get("stocks", []))

            rating = product["rating"]
            feedbacks = product["feedbacks"]

            extended = product.get("extended")
            if extended is not None:
                basic_sale = extended.get("basicSale", 0)

                basic_price_u = extended.get("basicPriceU")
                if basic_price_u is not None:
                    basic_price_u = basic_price_u / 100
                else:
                    basic_price_u = price_u

                client_sale = extended["clientSale"]
                client_price_u = extended["clientPriceU"]/100
            else:
                basic_sale = ""
                basic_price_u = ""
                client_sale = ""
                client_price_u = ""

        else:
            sku = ""
            name = ""
            price_u = ""
            basic_sale = ""
            basic_price_u = ""
            client_sale = ""
            client_price_u = ""
            qty = ""
            rating = ""
            feedbacks = ""

        # Записываем результаты в excel-файл
        row_values = [sku, name, price_u, basic_sale, basic_price_u,
                      client_sale, client_price_u, qty, rating, feedbacks]
        for i, value in enumerate(row_values):
            sheet.cell(row=row, column=i+1, value=value)

        print(f"\rSKU #{row-1} {sku} ok", end="")
        row += 1

# Сохраняем excel-файл
wb.save("Отчет по скидкам.xlsx")
print()
print("Made by https://t.me/ArChernushevich")
input("Нажмите Enter, чтобы выйти...")
