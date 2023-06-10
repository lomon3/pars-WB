# pip install pandas
# pip install requests
# pip install pyinstaller

import requests
import openpyxl

# Получаем значение spp
spp = 20
'''
try:
    response = requests.get("http://10.178.2.65/spp/")
    spp = response.json()["spp"]
except:
    print("Failed to get spp from server")'''

# Читаем excel-файл
wb = openpyxl.load_workbook("Отчет по скидкам.xlsx")
sheet = wb.active

# Добавляем новые столбцы
sheet.insert_cols(2, 6)
sheet.cell(row=1, column=2, value="priceU")
sheet.cell(row=1, column=3, value="basicSale")
sheet.cell(row=1, column=4, value="basicPriceU")
sheet.cell(row=1, column=5, value="clientSale")
sheet.cell(row=1, column=6, value="clientPriceU")
sheet.cell(row=1, column=7, value="qty")

# Обходим строки с артикулами
for row in range(2, sheet.max_row + 1):
    art = str(sheet.cell(row=row, column=1).value)

    # Формируем ссылку для запроса
    url = f"https://card.wb.ru/cards/detail?spp={spp}&reg=1&locale=ru&dest=-1216601,-115136,-421732,123585595&nm={art}"

    # Отправляем запрос и получаем ответ в формате JSON
    try:
        response = requests.get(url)
        data = response.json()
    except Exception as e:
        print(f"Failed to get data for art= {art}: {e}")
        continue

    # Получаем значения из ответа
    product = data["data"]["products"]
    if product:

        extended = product[0]["extended"]
        price_u = product[0]["priceU"]/100

        basic_sale = extended.get("basicSale", 0)
        
        basic_price_u = extended.get("basicPriceU")
        if basic_price_u is not None:
            basic_price_u = basic_price_u / 100
        else:
            basic_price_u = price_u

        client_sale = extended["clientSale"]
        client_price_u = extended["clientPriceU"]/100

        sizes = product[0].get("sizes", [])
        qty = sum(stock["qty"]
                  for size in sizes for stock in size.get("stocks", []))

    else:
        price_u = ""
        basic_sale = ""
        basic_price_u = ""
        client_sale = ""
        client_price_u = ""
        qty = ""

    # Записываем результаты в excel-файл
    row_values = [price_u, basic_sale, basic_price_u,
                  client_sale, client_price_u, qty]
    for i, value in enumerate(row_values):
        sheet.cell(row=row, column=i+2, value=value)

    print(f"\rSKU #{row-1} {art} ok", end="")

# Сохраняем excel-файл
wb.save("Отчет по скидкам.xlsx")
print()
input("Нажмите Enter, чтобы выйти...")