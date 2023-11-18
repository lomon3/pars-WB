import requests
import aiohttp
import asyncio
from openpyxl import load_workbook

# Получаем значение spp
def get_spp(cookie_value):
    headers = {'cookie': cookie_value}
    url = 'https://www.wildberries.ru/webapi/personalinfo/extrainfo'
    while True:
        response = requests.post(url, headers=headers)
        if response.status_code == 200:
            personal_discount = response.json()['value']['discountInfo']['personalDiscount']
            return personal_discount
        else:
            print(f"Ошибка запроса: {response.status_code}. Повторная попытка...")

# Асинхронная функция для отправки запросов
async def fetch(session, url):
    try:
        async with session.get(url) as response:
            return await response.json()
    except Exception as e:
        print(f"Ошибка при запросе к {url}: {e}")
        return None

# Асинхронная функция для обработки частей списка артикулов

async def process_article_chunks(article_chunks, spp, sheet):
    async with aiohttp.ClientSession() as session:
        tasks = []
        for chunk in article_chunks:
            url = f"https://card.wb.ru/cards/v1/detail?appType=0&curr=rub&dest=-1257484&spp={spp}&nm={';'.join(chunk)}"
            tasks.append(fetch(session, url))

        responses = await asyncio.gather(*tasks)

        # Создаем словарь для хранения данных
        products_data = {}

        # Обработка ответов
        for response in responses:
            if response:
                products = response["data"]["products"]
                for product in products:
                    sku = product.get("id", "")
                    name = product.get("name", "")
                    price_u = product.get("priceU", 0) / 100

                    extended = product.get("extended")
                    if extended is not None:
                        basic_sale = extended.get("basicSale", 0)

                        basic_price_u = extended.get("basicPriceU")
                        if basic_price_u is not None:
                            basic_price_u = basic_price_u / 100
                        else:
                            basic_price_u = price_u

                        client_sale = extended.get("clientSale", 0)

                        client_price_u = extended.get("clientPriceU")
                        if client_price_u is not None:
                            client_price_u = client_price_u / 100
                        else:
                            client_price_u = basic_price_u

                    else:
                        basic_sale = ""
                        basic_price_u = ""
                        client_sale = ""
                        client_price_u = product.get("salePriceU", 0) / 100

                    qty = sum(stock["qty"] for size in product.get("sizes", []) for stock in size.get("stocks", []))
                    rating = product.get("rating", 0)
                    feedbacks = product.get("feedbacks", 0)
                    subjectId = product.get("subjectId", 0)
                    subjectParentId = product.get("subjectParentId", 0)

                    products_data[sku] = [name, price_u, basic_sale, basic_price_u, client_sale, client_price_u, qty, rating, feedbacks, subjectId, subjectParentId]
                    print(f"{sku}: {products_data[sku]}")

        # Обновляем данные в Excel на основе словаря
        for row in sheet.iter_rows(min_row=2):
            sku = row[2].value
            if sku in products_data:
                for i, value in enumerate(products_data[sku]):
                    row[i + 12].value = value


# Главная функция
async def main():
    wb = load_workbook("Шаблон обновления цен и скидок.xlsx")
    sheet = wb['Отчет - цены и скидки на товары']
    
    try:
        with open("config.txt", "r") as file:
            client_cookie = file.read().strip()
    except FileNotFoundError:
        print("Файл 'config.txt' не найден в текущей директории.")
        input("Убедитесь, что файл находится в той же папке, что и скрипт.")
        return  # Завершаем выполнение функции main

    spp = input("Введите СПП или оставьте пустым для автоматической загрузки СПП: ") or get_spp(client_cookie)
    print(f'Сегодня СПП = {spp}%')

    sheet.cell(row=1, column=13, value="Название")
    sheet.cell(row=1, column=14, value="Цена до ВСЕХ скидок")
    sheet.cell(row=1, column=15, value="Скидка поставщика")
    sheet.cell(row=1, column=16, value="Цена поставщика")
    sheet.cell(row=1, column=17, value="СПП")
    sheet.cell(row=1, column=18, value="Цена на сайте")
    sheet.cell(row=1, column=19, value="сток")
    sheet.cell(row=1, column=20, value="рейт")
    sheet.cell(row=1, column=21, value="отзывы")
    sheet.cell(row=1, column=22, value="subjectId")
    sheet.cell(row=1, column=23, value="subjectParentId")

    # Логика группировки списка и определения чанков
    chunk_size = 300
    art_list = []
    for row in range(2, sheet.max_row + 1):
        art = str(sheet.cell(row=row, column=3).value)
        art_list.append(art)
    article_chunks = [art_list[i:i+chunk_size] for i in range(0, len(art_list), chunk_size)]

    await process_article_chunks(article_chunks, spp, sheet)

    success = False
    while not success:
        try:
            print("Сохраняю файл...")
            wb.save("Шаблон обновления цен и скидок.xlsx")
            print("Обработка завершена")
            success = True
        
        except PermissionError as e:
            print(f"Ошибка доступа к файлу Шаблон обновления цен и скидок.xlsx. Пожалуйста, закройте файл и нажмите Enter для повторной попытки, или введите 'ggwp' для отмены. Ошибка: {e}")
            response = input().strip().lower()
            if response == 'ggwp':
                print("Операция отменена пользователем.")
                break
        except Exception as e:
            print(f"Произошла неизвестная ошибка при обработке файла: {e}")
            break

    print()
    print("Made by https://t.me/ArChernushevich")
    input("Нажмите Enter, чтобы выйти...")

if __name__ == "__main__":
    asyncio.run(main())
