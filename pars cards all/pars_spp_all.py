
import requests
from openpyxl import load_workbook

# Получаем значение spp
def get_spp():
    headers = {
    'cookie': cookie_value,
    }

    url = 'https://www.wildberries.ru/webapi/personalinfo/extrainfo'

    while True:
        response = requests.post(url, headers=headers)
        if response.status_code == 200:
            personal_discount = response.json()['value']['discountInfo']['personalDiscount']
            return personal_discount
        else:
            print(f"Ошибка запроса: {response.status_code}. Повторная попытка...")

# Открыть книгу Excel и получить листы
wb = load_workbook("Отчет по скидкам.xlsx")
ws = wb["cookie"]
cookie_value = ws["A1"].value
sheet = wb['Общий отчет']

spp = input("Введите СПП или оставьте пустым для автоматической загрузки СПП: ") or get_spp()
print(f'Сегодня СПП = {spp}%')

# Добавляем новые столбцы
sheet.cell(row=1, column=2, value="Арт ВБ")
sheet.cell(row=1, column=3, value="Название")
sheet.cell(row=1, column=4, value="Цена до скидок")
sheet.cell(row=1, column=5, value="Скидка поставщика")
sheet.cell(row=1, column=6, value="Цена после скидки поставщика")
sheet.cell(row=1, column=7, value="СПП")
sheet.cell(row=1, column=8, value="Цена после СПП")
sheet.cell(row=1, column=9, value="сток")
sheet.cell(row=1, column=10, value="рейт")
sheet.cell(row=1, column=11, value="отзывы")
sheet.cell(row=1, column=12, value="subjectId")
sheet.cell(row=1, column=13, value="subjectParentId")

# группируем список
chunk_size = 300
art_list = []
for row in range(2, sheet.max_row + 1):
    art = str(sheet.cell(row=row, column=1).value)
    art_list.append(art)
    article_chunks = [art_list[i:i+chunk_size]
                      for i in range(0, len(art_list), chunk_size)]

row = 2
for chunk in article_chunks:
    url = f"https://card.wb.ru/cards/detail?spp={spp}&reg=1&locale=ru&dest=-1216601,-115136,-421732,123585595&nm={';'.join(chunk)}"
    # print(url)

    # Отправляем запрос и получаем ответ в формате JSON
    try:
        response = requests.get(url)
        data = response.json()
    except Exception as e:
        print(f"Failed to get data for art= {chunk}: {e}")
        continue
    products = data["data"]["products"]
    for product in products:
        if product:

            sku = product["id"]
            name = product["name"]

            price_u = product.get("priceU")
            if price_u is not None:
                price_u = price_u / 100

            sizes = product.get("sizes", [])
            qty = sum(stock["qty"]
                      for size in sizes for stock in size.get("stocks", []))

            rating = product["rating"]
            feedbacks = product["feedbacks"]
            subjectId = product["subjectId"]
            subjectParentId = product["subjectParentId"]

            extended = product.get("extended")
            if extended is not None:
                basic_sale = extended.get("basicSale", 0)

                basic_price_u = extended.get("basicPriceU")
                if basic_price_u is not None:
                    basic_price_u = basic_price_u / 100
                else:
                    basic_price_u = price_u

                client_sale = extended.get("clientSale", 0)

                client_price_u = extended.get("clientPriceU")
                if client_price_u is not None:
                    client_price_u = client_price_u / 100
                else:
                    client_price_u = basic_price_u

            else:
                basic_sale = ""
                basic_price_u = ""
                client_sale = ""
                client_price_u = ""

        else:
            sku = ""
            name = ""
            price_u = ""
            basic_sale = ""
            basic_price_u = ""
            client_sale = ""
            client_price_u = ""
            qty = ""
            rating = ""
            feedbacks = ""
            subjectId = ""
            subjectParentId = ""

        # Записываем результаты в excel-файл
        row_values = [sku, name, price_u, basic_sale, basic_price_u,
                      client_sale, client_price_u, qty, rating, feedbacks, subjectId, subjectParentId]
        for i, value in enumerate(row_values):
            sheet.cell(row=row, column=i+2, value=value)

        print(f"\rSKU #{row-1} {sku} ok", end="")
        row += 1

# Сохраняем excel-файл
wb.save("Отчет по скидкам.xlsx")
print()
print("Made by https://t.me/ArChernushevich")
input("Нажмите Enter, чтобы выйти...")