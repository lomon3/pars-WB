import http.client
import json
import requests
import openpyxl
import os
import glob

def select_file():
    # Получение текущего рабочего каталога
    current_dir = os.getcwd()

    # Шаблон имени файла для поиска (например, '*.xlsx')
    file_pattern = '*.xlsx'

    # Полный путь к папке с файлами Excel
    folder_path = os.path.join(current_dir, file_pattern)

    # Получение списка файлов Excel в папке
    excel_files = glob.glob(folder_path)

    # Вывод списка файлов с порядковыми номерами
    for i, file_name in enumerate(excel_files):
        print(f"{i+1}. {os.path.basename(file_name)}")
    print("0. обработать все файлы")
    print()

    # Пользователь выбирает номер файла для обработки
    selected_file_num = int(input("Введите номер файла для обработки: "))
    print()

    # Проверка выбранного номера файла
    if 1 <= selected_file_num <= len(excel_files):
        selected_file_name = excel_files[selected_file_num - 1]
        # Здесь можете обработать выбранный файл
        print(f"Выбран файл: {os.path.basename(selected_file_name)}")
        print()
        return [selected_file_name]
    else:
        print("Выбраны все файлы")
        return excel_files


def get_position(conn, headers, company_id, query, item_ids):
    # Проверка наличия корректных item_id
    item_ids = [item_id for item_id in item_ids if item_id.isdigit()]
    #60617 204063
    payload = json.dumps({
        "company_id": company_id,
        "is_legal": False,
        "query": query,
        "location_uid": "0c5b2444-70a0-4932-980c-b4dc0d3f02b5",
        "item_ids": item_ids
    })

    conn.request("POST", "/api/validator-service/v1/get_search_stats", payload, headers)
    res = conn.getresponse()
    data = res.read()

    response_json = json.loads(data.decode("utf-8"))

    items = response_json.get('items', [])

    result = {}
    if items:
        for item in items:
            item_id = item['itemId']
            position = item['position']
            search_status = item['searchStatus']

            if search_status == 'FOUND' and position != '0':
                position = int(position)
            elif search_status == 'FOUND' and position == '0':
                position = 1500
            elif search_status == 'NOT_FOUND' and position == '0':
                position = 0

            result[item_id] = position

    return result


def pars_selected_files(selected_file):

    conn = http.client.HTTPSConnection("seller.ozon.ru")

    print(selected_file)
    # Открыть книгу Excel и получить листы
    ozon_file = openpyxl.load_workbook(selected_file)
    sheet = ozon_file['ozon']
    max_col = sheet.max_column

    cookie = str(ozon_file['cookie'].cell(row=1, column=1).value)
    company_id = str(ozon_file['cookie'].cell(row=5, column=1).value)
    headers = {
        'Cookie': cookie,
        'Content-Type': 'application/json'
    }

    for row in sheet.iter_rows(min_row=3, min_col=4, max_row=sheet.max_row, max_col=max_col):
        for cell in row:
            cell.value = ""
    print('sheet cleaned')

    # Индексированный словарь для хранения значений столбца 2
    keywords = {row[0].row: str(row[1].value) for row in sheet.iter_rows(
        min_row=3, max_row=sheet.max_row)}
    
    art_list = []
    for col in range(4, max_col + 1):
      art = str(sheet.cell(row=2, column=col).value)
      art_list.append(art)

    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
        row_num = row[0].row
        keyword = keywords[row_num]
        pos = get_position(conn, headers, company_id, keyword, art_list)
        
        cell_values = [str(sheet.cell(row=2, column=col).value) for col in range(4, max_col + 1)]
        
        for col, cell_value in enumerate(cell_values, start=4):
            position = pos.get(cell_value)
            if position is not None:
                sheet.cell(row=row_num, column=col, value=position)
        
        print(f"\rquery: {keyword} - ok", end="")

    # Сохраняем excel-файл
    ozon_file.save(selected_file)
    print()


selected_files = []
selected_files += select_file()
for files in selected_files:
    pars_selected_files(files)

print("Made by https://t.me/ArChernushevich")
input("Нажмите Enter, чтобы выйти...")